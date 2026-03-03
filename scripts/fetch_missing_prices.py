#!/usr/bin/env python3
"""
通过 akshare 行情API补齐缺失的买卖价格，并生成最终合并Excel

输入:
  - output/tgb_zhihedaxuesheng_买卖记录.json   (买卖记录, 含股票代码)
  - output/tgb_zhihedaxuesheng_OCR持仓数据.json (OCR数据, 含成本价/现价)

输出:
  - output/tgb_只核大学生_交易价格明细_补齐.xlsx  (最终合并Excel)
  - output/tgb_zhihedaxuesheng_行情价格缓存.json (API结果缓存)
"""
import json
import os
import time
import re

# 绕过系统代理，直连东方财富行情接口
os.environ['NO_PROXY'] = '*'
os.environ['no_proxy'] = '*'
for key in ['http_proxy', 'https_proxy', 'HTTP_PROXY', 'HTTPS_PROXY', 'ALL_PROXY', 'all_proxy']:
    os.environ.pop(key, None)

import requests
# 强制 requests 忽略系统代理设置（包括 macOS 系统级代理）
_original_init = requests.Session.__init__
def _patched_init(self, *args, **kwargs):
    _original_init(self, *args, **kwargs)
    self.trust_env = False
    self.proxies = {'http': None, 'https': None}
requests.Session.__init__ = _patched_init

import pandas as pd
import akshare as ak

OUTPUT_DIR = "/Users/tq/PycharmProjects/stocks_analysis/output"
TRADES_JSON = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_买卖记录.json")
OCR_JSON = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_OCR持仓数据.json")
CACHE_FILE = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_行情价格缓存.json")
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, "tgb_只核大学生_交易价格明细_补齐.xlsx")

# OCR 噪音名称（与 ocr_tgb_screenshots.py 保持一致）
NOISE_NAMES = {
    "持仓管理", "持仓资讯", "持仓分时", "市值今", "市值合",
    "白立併日", "涵行壮夂", "合答理",
    "淘阳吧", "淘限吧", "淘胞吧", "淘服吧", "淘股吧", "淘肪",
    "淘脂吧", "淘批量买入", "海限吧", "海頭吧", "海灣過",
    "海我滥买入", "海級翻可台", "渝肪", "密肪",
    "溷胞", "溜股", "溷股", "酱灣", "图股",
    "汽量买入", "汽批量实入", "方修淘股吧",
    "淘特仓资讯", "行情愛",
}

# 特殊名称映射（OCR名称 → 买卖记录名称，或反向）
SPECIAL_NAME_MAP = {
    "朗新集团": "朗新科技",
    "朗新科技": "朗新集团",
}

# 行情缓存
_price_cache = {}


def is_noise_name(name):
    """判断是否为OCR噪音"""
    if not isinstance(name, str) or not name.strip():
        return True
    name = name.strip()
    if name in NOISE_NAMES:
        return True
    if "淘" in name and "吧" in name:
        return True
    if "海" in name and "吧" in name:
        return True
    ui_keywords = ["持仓", "管理", "批量", "行情", "挂合", "市值"]
    for kw in ui_keywords:
        if kw in name and len(name) <= 5:
            return True
    return False


def load_cache():
    """加载API结果缓存"""
    global _price_cache
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            _price_cache = json.load(f)
        print(f"  已加载 {len(_price_cache)} 条价格缓存")


def save_cache():
    """保存API结果缓存"""
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(_price_cache, f, ensure_ascii=False, indent=2)


def fetch_stock_price(code, date_str):
    """
    获取股票在指定日期的收盘价
    code: 6位股票代码, 如 "002449"
    date_str: "2025-01-02" 或 "20250102"
    返回: float收盘价 或 None
    """
    date_clean = date_str.replace("-", "")
    cache_key = f"{code}_{date_clean}"

    if cache_key in _price_cache:
        return _price_cache[cache_key]

    try:
        # akshare 日K线接口
        df = ak.stock_zh_a_hist(
            symbol=code,
            period="daily",
            start_date=date_clean,
            end_date=date_clean,
            adjust="qfq"
        )
        if df is not None and len(df) > 0:
            close = float(df.iloc[0]["收盘"])
            _price_cache[cache_key] = close
            save_cache()
            time.sleep(0.3)
            return close

        # 当天无数据（非交易日），尝试前后5天
        target = pd.Timestamp(date_clean)
        start = (target - pd.Timedelta(days=5)).strftime("%Y%m%d")
        end = (target + pd.Timedelta(days=5)).strftime("%Y%m%d")
        df = ak.stock_zh_a_hist(
            symbol=code, period="daily",
            start_date=start, end_date=end, adjust="qfq"
        )
        if df is not None and len(df) > 0:
            df["日期"] = pd.to_datetime(df["日期"])
            df["距离"] = abs(df["日期"] - target)
            closest = df.loc[df["距离"].idxmin()]
            close = float(closest["收盘"])
            _price_cache[cache_key] = close
            save_cache()
            time.sleep(0.3)
            return close

    except Exception as e:
        print(f"    获取行情失败 {code} {date_str}: {e}")

    time.sleep(0.3)
    return None


def clean_ocr_data(ocr_data):
    """离线清洗OCR数据（过滤噪音、修正异常值）"""
    cleaned_count = 0
    fixed_pct_count = 0
    fixed_shares_count = 0

    for date_str, entry in ocr_data.items():
        original = entry.get("holdings", [])
        cleaned = []
        for h in original:
            name = h.get("股票名称", "")
            if is_noise_name(name):
                cleaned_count += 1
                continue

            cost = h.get("成本价")
            price = h.get("现价")

            # 交叉验证盈亏百分比
            if cost and price and cost > 0:
                calculated_pct = round((price - cost) / cost * 100, 2)
                ocr_pct = h.get("盈亏百分比")
                if ocr_pct is None or abs(ocr_pct - calculated_pct) > 1.0:
                    h["盈亏百分比"] = calculated_pct
                    fixed_pct_count += 1

            # 交叉验证持仓数量
            shares = h.get("持仓数量")
            market_val = h.get("市值")
            if price and price > 0 and market_val and market_val > 0:
                expected_shares = round(market_val / price / 100) * 100
                if expected_shares > 0 and shares:
                    deviation = abs(shares - expected_shares) / expected_shares
                    if deviation > 0.5:
                        h["持仓数量"] = expected_shares
                        fixed_shares_count += 1

            cleaned.append(h)

        entry["holdings"] = cleaned

    print(f"  OCR清洗: 过滤噪音 {cleaned_count} 条, 修正盈亏% {fixed_pct_count} 条, 修正持仓数量 {fixed_shares_count} 条")
    return ocr_data


def build_ocr_index(ocr_data):
    """构建OCR数据索引，支持按名称和别名查找"""
    # (date, name) -> holding_info
    index = {}
    for date_str, entry in ocr_data.items():
        for h in entry.get("holdings", []):
            name = h.get("股票名称", "")
            if not name:
                continue
            index[(date_str, name)] = h
    return index


def find_ocr_price(ocr_index, date_str, stock_name, field):
    """
    从OCR索引中查找价格
    field: "成本价" 或 "现价"
    支持名称别名查找
    """
    # 直接匹配
    key = (date_str, stock_name)
    if key in ocr_index:
        val = ocr_index[key].get(field)
        if val is not None:
            return val, "OCR"

    # 别名匹配
    alt_name = SPECIAL_NAME_MAP.get(stock_name)
    if alt_name:
        key_alt = (date_str, alt_name)
        if key_alt in ocr_index:
            val = ocr_index[key_alt].get(field)
            if val is not None:
                return val, "OCR"

    return None, None


def get_trading_dates(ocr_data):
    """从OCR数据获取所有交易日期，并排序"""
    dates = sorted(ocr_data.keys())
    return dates


def get_prev_trading_date(trading_dates, date_str):
    """获取前一个交易日"""
    date_clean = date_str.replace("-", "")
    for i, d in enumerate(trading_dates):
        if d == date_clean and i > 0:
            return trading_dates[i - 1]
    # 如果精确匹配不到，找最近的前一天
    for i, d in enumerate(trading_dates):
        if d >= date_clean and i > 0:
            return trading_dates[i - 1]
    return None


def match_prices(trade_records, ocr_data):
    """
    匹配买卖价格
    - 买入价: 买入当日OCR成本价
    - 卖出价: 卖出前一交易日OCR现价（因为卖出当天截图中已无该股票）
    - 缺失的通过akshare API补齐
    """
    ocr_index = build_ocr_index(ocr_data)
    trading_dates = get_trading_dates(ocr_data)

    results = []
    stats = {"buy_ocr": 0, "buy_api": 0, "buy_miss": 0,
             "sell_ocr": 0, "sell_api": 0, "sell_miss": 0}

    for t in trade_records:
        date_str = str(t["date"])
        date_fmt = t["date_str"]
        stock_name = t.get("name", "")
        code = t.get("code", "")
        action = t["action"]

        rec = {
            "action": action,
            "date": t["date"],
            "date_str": date_fmt,
            "code": code,
            "name": stock_name,
            "buy_date": t.get("buy_date", ""),
            "hold_days": t.get("hold_days", ""),
            "asset_after": t.get("asset_after", ""),
            "day_return": t.get("day_return", ""),
            "total_return": t.get("total_return", ""),
            "position": t.get("position", ""),
            "buy_price": None,
            "sell_price": None,
            "price_source": "",
        }

        if action == "买入":
            # 买入价 = OCR成本价（买入当日）
            price, source = find_ocr_price(ocr_index, date_str, stock_name, "成本价")
            if price is not None:
                rec["buy_price"] = price
                rec["price_source"] = "OCR成本价"
                stats["buy_ocr"] += 1
            else:
                # API补齐
                if code:
                    print(f"  API补齐买入价: {stock_name}({code}) {date_fmt}")
                    price = fetch_stock_price(code, date_fmt)
                    if price:
                        rec["buy_price"] = round(price, 3)
                        rec["price_source"] = "行情API收盘价"
                        stats["buy_api"] += 1
                    else:
                        rec["price_source"] = "缺失"
                        stats["buy_miss"] += 1
                else:
                    rec["price_source"] = "缺失"
                    stats["buy_miss"] += 1

        elif action == "卖出":
            # 卖出价 = 前一交易日OCR现价
            prev_date = get_prev_trading_date(trading_dates, date_fmt)
            if prev_date:
                price, source = find_ocr_price(ocr_index, prev_date, stock_name, "现价")
                if price is not None:
                    rec["sell_price"] = price
                    rec["price_source"] = "OCR现价(前日)"
                    stats["sell_ocr"] += 1
                else:
                    # 也尝试当日OCR（部分截图可能在盘中截取）
                    price, source = find_ocr_price(ocr_index, date_str, stock_name, "现价")
                    if price is not None:
                        rec["sell_price"] = price
                        rec["price_source"] = "OCR现价(当日)"
                        stats["sell_ocr"] += 1
                    elif code:
                        # API补齐
                        print(f"  API补齐卖出价: {stock_name}({code}) {date_fmt}")
                        price = fetch_stock_price(code, date_fmt)
                        if price:
                            rec["sell_price"] = round(price, 3)
                            rec["price_source"] = "行情API收盘价"
                            stats["sell_api"] += 1
                        else:
                            rec["price_source"] = "缺失"
                            stats["sell_miss"] += 1
                    else:
                        rec["price_source"] = "缺失"
                        stats["sell_miss"] += 1
            elif code:
                # 无前一交易日数据，直接API
                print(f"  API补齐卖出价: {stock_name}({code}) {date_fmt}")
                price = fetch_stock_price(code, date_fmt)
                if price:
                    rec["sell_price"] = round(price, 3)
                    rec["price_source"] = "行情API收盘价"
                    stats["sell_api"] += 1
                else:
                    rec["price_source"] = "缺失"
                    stats["sell_miss"] += 1
            else:
                rec["price_source"] = "缺失"
                stats["sell_miss"] += 1

            # 为卖出记录补充买入价（从对应的买入记录中查找）
            buy_date = t.get("buy_date", "")
            if buy_date:
                buy_date_clean = buy_date.replace("-", "")
                price, source = find_ocr_price(ocr_index, buy_date_clean, stock_name, "成本价")
                if price is not None:
                    rec["buy_price"] = price

        results.append(rec)

    print(f"\n  匹配统计:")
    print(f"    买入价: OCR {stats['buy_ocr']}, API {stats['buy_api']}, 缺失 {stats['buy_miss']}")
    print(f"    卖出价: OCR {stats['sell_ocr']}, API {stats['sell_api']}, 缺失 {stats['sell_miss']}")

    return results, stats


def generate_excel(results, stats):
    """生成最终合并Excel，标注价格来源"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10)
    green_font = Font(name="微软雅黑", size=10, color="008000", bold=True)
    red_font = Font(name="微软雅黑", size=10, color="FF0000", bold=True)
    title_font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    buy_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    sell_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    api_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    missing_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    # ========== Sheet 1: 交易价格明细 ==========
    ws1 = wb.active
    ws1.title = "交易价格明细(补齐版)"

    ws1.merge_cells('A1:M1')
    ws1['A1'] = "淘股吧 2025梦想杯 - 只核大学生 交易价格明细（OCR + 行情API补齐）"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 35

    headers = [
        "操作", "日期", "股票代码", "股票名称", "买入日期", "持仓天数",
        "买入价", "卖出价", "单笔盈亏(%)",
        "操作后总资产(万)", "当日收益(%)", "累计收益(%)", "价格来源"
    ]
    col_widths = [8, 14, 12, 12, 14, 10, 12, 12, 12, 14, 12, 12, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[2].height = 30

    # 按时间倒序
    sorted_results = sorted(results, key=lambda x: x["date"], reverse=True)

    row_idx = 3
    for rec in sorted_results:
        is_buy = rec["action"] == "买入"
        buy_price = rec.get("buy_price")
        sell_price = rec.get("sell_price")

        # 计算单笔盈亏%
        trade_pnl = None
        if rec["action"] == "卖出" and buy_price and sell_price and buy_price > 0:
            trade_pnl = round((sell_price - buy_price) / buy_price * 100, 2)

        row_data = [
            rec["action"],
            rec["date_str"],
            rec["code"],
            rec["name"],
            rec.get("buy_date", ""),
            rec.get("hold_days", ""),
            buy_price,
            sell_price,
            trade_pnl,
            rec.get("asset_after", ""),
            rec.get("day_return", ""),
            rec.get("total_return", ""),
            rec.get("price_source", ""),
        ]

        row_bg = buy_fill if is_buy else sell_fill
        source = rec.get("price_source", "")

        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = row_bg

            if col_idx <= 6:
                cell.alignment = center_align
            else:
                cell.alignment = right_align

            if col_idx == 1:
                cell.font = green_font if is_buy else red_font
            elif col_idx == 9 and val is not None:
                # 盈亏颜色
                try:
                    v = float(val)
                    cell.font = Font(name="微软雅黑", size=10, color="008000" if v > 0 else "FF0000" if v < 0 else "000000")
                except (ValueError, TypeError):
                    cell.font = data_font
            else:
                cell.font = data_font

            # 价格来源特殊标色
            if col_idx in (7, 8, 13) and "行情API" in source:
                cell.fill = api_fill
            elif col_idx in (7, 8, 13) and source == "缺失":
                cell.fill = missing_fill

        row_idx += 1

    ws1.freeze_panes = "A3"

    # ========== Sheet 2: 完整交易配对 ==========
    ws2 = wb.create_sheet("完整交易配对")

    ws2.merge_cells('A1:L1')
    ws2['A1'] = "已完成交易配对（买入→卖出）"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 35

    headers2 = [
        "股票代码", "股票名称", "买入日期", "卖出日期", "持仓天数",
        "买入价", "卖出价", "盈亏(%)", "盈亏金额(估)",
        "买入价来源", "卖出价来源", "总资产(万)"
    ]
    col_widths2 = [12, 12, 14, 14, 10, 12, 12, 12, 14, 14, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[2].height = 30

    # 构建买入记录索引: (code, buy_date) -> rec
    buy_index = {}
    for rec in results:
        if rec["action"] == "买入":
            key = (rec["code"], rec["date_str"])
            buy_index[key] = rec

    # 遍历卖出记录，配对买入
    sell_records = [r for r in results if r["action"] == "卖出"]
    sell_records.sort(key=lambda x: x["date"], reverse=True)

    row_idx = 3
    total_trades = 0
    profitable_trades = 0

    for sell_rec in sell_records:
        buy_date = sell_rec.get("buy_date", "")
        code = sell_rec["code"]

        # 查找对应买入记录
        buy_rec = buy_index.get((code, buy_date), {})
        buy_price = buy_rec.get("buy_price") or sell_rec.get("buy_price")
        buy_source = buy_rec.get("price_source", "")
        sell_price = sell_rec.get("sell_price")
        sell_source = sell_rec.get("price_source", "")

        # 盈亏计算
        pnl_pct = None
        pnl_amount = None
        if buy_price and sell_price and buy_price > 0:
            pnl_pct = round((sell_price - buy_price) / buy_price * 100, 2)
            # 估算盈亏金额（假设10万本金满仓）
            pnl_amount = round((sell_price - buy_price) / buy_price * 100000, 0)

        total_trades += 1
        if pnl_pct is not None and pnl_pct > 0:
            profitable_trades += 1

        row_data = [
            code,
            sell_rec["name"],
            buy_date,
            sell_rec["date_str"],
            sell_rec.get("hold_days", ""),
            buy_price,
            sell_price,
            pnl_pct,
            pnl_amount,
            buy_source,
            sell_source,
            sell_rec.get("asset_after", ""),
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = data_font
            if col_idx <= 5:
                cell.alignment = center_align
            else:
                cell.alignment = right_align

            # 盈亏颜色
            if col_idx == 8 and val is not None:
                try:
                    v = float(val)
                    cell.font = Font(name="微软雅黑", size=10, bold=True,
                                     color="008000" if v > 0 else "FF0000" if v < 0 else "000000")
                    cell.fill = PatternFill(start_color="E8F5E9" if v > 0 else "FFEBEE",
                                            end_color="E8F5E9" if v > 0 else "FFEBEE",
                                            fill_type="solid")
                except (ValueError, TypeError):
                    pass

            # 来源标色
            if col_idx in (10, 11) and "行情API" in str(val):
                cell.fill = api_fill

        row_idx += 1

    ws2.freeze_panes = "A3"

    # ========== Sheet 3: 数据源说明 ==========
    ws3 = wb.create_sheet("数据源说明")

    ws3.merge_cells('A1:D1')
    ws3['A1'] = "价格数据来源统计"
    ws3['A1'].font = title_font
    ws3['A1'].alignment = center_align
    ws3.row_dimensions[1].height = 35

    stat_headers = ["类别", "OCR匹配", "行情API补齐", "缺失"]
    for col_idx, header in enumerate(stat_headers, 1):
        cell = ws3.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 12

    # 买入价统计
    buy_total = stats["buy_ocr"] + stats["buy_api"] + stats["buy_miss"]
    for col_idx, val in enumerate(["买入价", stats["buy_ocr"], stats["buy_api"], stats["buy_miss"]], 1):
        cell = ws3.cell(row=3, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = center_align

    # 卖出价统计
    sell_total = stats["sell_ocr"] + stats["sell_api"] + stats["sell_miss"]
    for col_idx, val in enumerate(["卖出价", stats["sell_ocr"], stats["sell_api"], stats["sell_miss"]], 1):
        cell = ws3.cell(row=4, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = center_align

    # 汇总行
    ws3.cell(row=6, column=1, value="交易配对统计").font = Font(name="微软雅黑", bold=True, size=11)
    ws3.cell(row=7, column=1, value="总完成交易").font = data_font
    ws3.cell(row=7, column=2, value=total_trades).font = data_font
    ws3.cell(row=8, column=1, value="盈利交易").font = data_font
    ws3.cell(row=8, column=2, value=profitable_trades).font = data_font
    if total_trades > 0:
        win_rate = round(profitable_trades / total_trades * 100, 1)
        ws3.cell(row=9, column=1, value="胜率").font = data_font
        ws3.cell(row=9, column=2, value=f"{win_rate}%").font = Font(
            name="微软雅黑", size=10, bold=True,
            color="008000" if win_rate > 50 else "FF0000")

    ws3.cell(row=11, column=1, value="说明:").font = Font(name="微软雅黑", bold=True, size=10)
    ws3.cell(row=12, column=1, value="白色/绿色/红色背景 = OCR提取的价格").font = data_font
    ws3.cell(row=13, column=1, value="浅蓝色背景 = 行情API收盘价补齐").font = data_font
    ws3.cell(row=13, column=1).fill = api_fill
    ws3.cell(row=14, column=1, value="浅橙色背景 = 价格缺失").font = data_font
    ws3.cell(row=14, column=1).fill = missing_fill

    # 保存
    wb.save(OUTPUT_EXCEL)
    print(f"\n  Excel 已保存: {OUTPUT_EXCEL}")


def main():
    print("=" * 60)
    print("补齐交割单买卖价格（OCR + 行情API）")
    print("=" * 60)

    # 读取买卖记录
    print(f"\n  读取买卖记录: {TRADES_JSON}")
    with open(TRADES_JSON, "r", encoding="utf-8") as f:
        trade_records = json.load(f)
    print(f"  共 {len(trade_records)} 条记录")

    # 读取OCR数据
    print(f"  读取OCR数据: {OCR_JSON}")
    with open(OCR_JSON, "r", encoding="utf-8") as f:
        ocr_data = json.load(f)
    print(f"  共 {len(ocr_data)} 个交易日")

    # 加载API缓存
    load_cache()

    # 离线清洗OCR数据
    print("\n  清洗OCR数据...")
    ocr_data = clean_ocr_data(ocr_data)

    # 匹配价格
    print("\n  匹配买卖价格...")
    results, stats = match_prices(trade_records, ocr_data)

    # 生成Excel
    print("\n  生成Excel报告...")
    generate_excel(results, stats)

    # 最终统计
    buy_count = sum(1 for r in results if r["action"] == "买入")
    sell_count = sum(1 for r in results if r["action"] == "卖出")
    buy_with_price = sum(1 for r in results if r["action"] == "买入" and r.get("buy_price"))
    sell_with_price = sum(1 for r in results if r["action"] == "卖出" and r.get("sell_price"))

    print(f"\n{'='*60}")
    print(f"  补齐完成!")
    print(f"  买入: {buy_with_price}/{buy_count} 有价格 ({buy_with_price/max(buy_count,1)*100:.1f}%)")
    print(f"  卖出: {sell_with_price}/{sell_count} 有价格 ({sell_with_price/max(sell_count,1)*100:.1f}%)")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
