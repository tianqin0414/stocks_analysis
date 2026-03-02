#!/usr/bin/env python3
"""
将淘股吧交割单数据从 JSON/CSV 转换为格式化的 Excel 文件
"""
import json
import os
import sys

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

INPUT_JSON = "/Users/tq/PycharmProjects/stocks_analysis/output/tgb_zhihedaxuesheng_data.json"
OUTPUT_EXCEL = "/Users/tq/PycharmProjects/stocks_analysis/output/tgb_只核大学生_交割单.xlsx"


def stock_code_to_display(full_code):
    """将 sh600343 -> 600343"""
    if full_code and len(full_code) > 2:
        return full_code[2:]
    return full_code or ""


def format_date(date_num):
    """将 20251231 -> 2025-12-31"""
    s = str(date_num)
    if len(s) == 8:
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return s


def create_excel():
    print(f"📖 读取数据: {INPUT_JSON}")
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    income_records = data["income"]
    trades = data["trades"]
    
    wb = Workbook()
    
    # ============== 样式定义 ==============
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(name="微软雅黑", size=10)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_right = Alignment(horizontal="right", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center")
    
    green_font = Font(name="微软雅黑", size=10, color="008000")  # 正收益
    red_font = Font(name="微软雅黑", size=10, color="FF0000")    # 负收益
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    
    title_font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="微软雅黑", size=10, color="666666")
    
    # ============== Sheet 1: 收益明细 ==============
    ws1 = wb.active
    ws1.title = "每日收益明细"
    
    # 标题行
    ws1.merge_cells('A1:O1')
    ws1['A1'] = "淘股吧 2025梦想杯 - 只核大学生 每日收益明细"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35
    
    ws1.merge_cells('A2:O2')
    ws1['A2'] = f"总收益: {income_records[0].get('totalRateD', 0)}% | 百万组排名第1 | 数据来源: tgb.cn | 下载时间: {data.get('downloadTime', '')}"
    ws1['A2'].font = subtitle_font
    ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 22
    
    # 表头
    headers1 = [
        "日期", "初始资产(万)", "昨日资产(万)", "当日资产(万)",
        "初始资产(元)", "昨日资产(元)", "当日资产(元)",
        "存取金额", "仓位(%)", "昨日收益(%)", "当日收益(%)", "总收益(%)",
        "排名", "持股数", "持股代码"
    ]
    col_widths1 = [14, 14, 14, 14, 14, 14, 14, 12, 10, 13, 13, 13, 8, 8, 50]
    
    for col_idx, (header, width) in enumerate(zip(headers1, col_widths1), 1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[3].height = 30
    
    # 数据行
    for row_idx, r in enumerate(income_records, 4):
        holdstocks = r.get("holdstocks") or ""
        if holdstocks:
            codes = [stock_code_to_display(c) for c in holdstocks.split(",")]
            holdstocks = ", ".join(codes)
        
        today_rate = r.get("todayRateD", 0) or 0
        
        row_data = [
            format_date(r.get("endDateNum")),
            r.get("firstMoneyStr"),
            r.get("preMoneyStr"),
            r.get("nowMoneyStr"),
            r.get("firstMoney"),
            r.get("preMoney"),
            r.get("nowMoney"),
            r.get("inoutMoneyStr", "0"),
            r.get("position"),
            r.get("preRateD"),
            today_rate,
            r.get("totalRateD"),
            r.get("sortNum"),
            r.get("holdStockNum"),
            holdstocks
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            if col_idx == 1:
                cell.alignment = data_align_center
            elif col_idx == 15:
                cell.alignment = data_align_left
            elif col_idx in (10, 11, 12):
                # 收益率列 - 根据正负设置颜色
                try:
                    v = float(val) if val else 0
                    if v > 0:
                        cell.font = green_font
                    elif v < 0:
                        cell.font = red_font
                except (ValueError, TypeError):
                    pass
                cell.alignment = data_align_right
            else:
                cell.alignment = data_align_right
            
            # 交替行背景色
            if (row_idx - 4) % 2 == 1:
                cell.fill = alt_fill
    
    # 冻结窗格
    ws1.freeze_panes = "A4"
    
    # ============== Sheet 2: 持仓明细 ==============
    ws2 = wb.create_sheet("每日持仓明细")
    
    ws2.merge_cells('A1:F1')
    ws2['A1'] = "淘股吧 2025梦想杯 - 只核大学生 每日持仓明细"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    
    headers2 = ["日期", "股票代码", "股票名称", "当日收益(%)", "金额", "数量"]
    col_widths2 = [14, 14, 14, 14, 14, 14]
    
    for col_idx, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[2].height = 30
    
    row_idx = 3
    sorted_dates = sorted(trades.keys(), reverse=True)
    for date_str in sorted_dates:
        trade_list = trades[date_str]
        for t in trade_list:
            code = stock_code_to_display(t.get("fullCode", ""))
            
            row_data = [
                format_date(date_str),
                code,
                t.get("stockName", ""),
                t.get("todayRate", ""),
                t.get("money", "--"),
                t.get("num", "--")
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = data_align_center
                
                if (row_idx - 3) % 2 == 1:
                    cell.fill = alt_fill
            
            row_idx += 1
    
    ws2.freeze_panes = "A3"
    
    # ============== Sheet 3: 持股汇总统计 ==============
    ws3 = wb.create_sheet("持股频次统计")
    
    # 统计每只股票出现的次数
    stock_freq = {}
    for date_str, trade_list in trades.items():
        for t in trade_list:
            code = stock_code_to_display(t.get("fullCode", ""))
            name = t.get("stockName", "")
            key = f"{code}_{name}"
            if key not in stock_freq:
                stock_freq[key] = {"code": code, "name": name, "count": 0, "dates": []}
            stock_freq[key]["count"] += 1
            stock_freq[key]["dates"].append(format_date(date_str))
    
    # 按频次排序
    sorted_stocks = sorted(stock_freq.values(), key=lambda x: x["count"], reverse=True)
    
    ws3.merge_cells('A1:E1')
    ws3['A1'] = "淘股吧 2025梦想杯 - 只核大学生 持股频次统计"
    ws3['A1'].font = title_font
    ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 35
    
    headers3 = ["排序", "股票代码", "股票名称", "持仓天数", "首次持仓日期"]
    col_widths3 = [8, 14, 14, 12, 16]
    
    for col_idx, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    ws3.row_dimensions[2].height = 30
    
    for row_idx, stock in enumerate(sorted_stocks, 3):
        first_date = min(stock["dates"]) if stock["dates"] else ""
        row_data = [
            row_idx - 2,
            stock["code"],
            stock["name"],
            stock["count"],
            first_date
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_align_center
            if (row_idx - 3) % 2 == 1:
                cell.fill = alt_fill
    
    ws3.freeze_panes = "A3"
    
    # ============== 保存 ==============
    print(f"💾 保存 Excel 文件: {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    
    print(f"\n✅ Excel 文件生成完成!")
    print(f"   📊 Sheet 1 '每日收益明细': {len(income_records)} 条记录")
    total_trades = sum(len(v) for v in trades.values())
    print(f"   📊 Sheet 2 '每日持仓明细': {total_trades} 条记录")
    print(f"   📊 Sheet 3 '持股频次统计': {len(sorted_stocks)} 只股票")
    print(f"   📁 文件路径: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    create_excel()
