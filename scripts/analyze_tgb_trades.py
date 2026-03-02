#!/usr/bin/env python3
"""
分析淘股吧交割单数据，推断买卖时间节点
通过对比每日持仓变化，推断买入和卖出的具体日期

同时，批量下载所有交割单截图（包含真实价格信息）
"""
import json
import os
import time
import urllib.request
from datetime import datetime
from collections import defaultdict

INPUT_JSON = "/Users/tq/PycharmProjects/stocks_analysis/output/tgb_zhihedaxuesheng_data.json"
OUTPUT_DIR = "/Users/tq/PycharmProjects/stocks_analysis/output"
IMAGES_DIR = os.path.join(OUTPUT_DIR, "tgb_交割单截图")
TRADES_JSON = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_买卖记录.json")
TRADES_CSV = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_买卖记录.csv")

# 用于下载截图的配置
USER_ID = "11310249"
SPMATCH_SEQ = "813"
COOKIES = (
    "tgbuser=12810600; "
    "tgbpwd=8daecd30e53af3098373334ba74c00c865a577cb782a854686a2e0d46979905dfpfqtq2qbqmq9v0; "
    "loginStatus=phone; "
    "JSESSIONID=NTY4ZmU0ZTEtZDNhZC00YWFjLWEzYmQtZmZhZTVhODQ4ZDNk; "
    "acw_tc=7929ee2c17723648343801290ee26e79d8a7e27c90e4189f49d27b800ea160"
)
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": f"https://www.tgb.cn/spmatch/gains/readInfo?lookeUserID={USER_ID}&spmatchSeq={SPMATCH_SEQ}",
    "Cookie": COOKIES,
}

# 股票代码到名称的映射（从持仓数据中收集）
STOCK_NAME_MAP = {}


def stock_code_to_display(full_code):
    if full_code and len(full_code) > 2:
        return full_code[2:]
    return full_code or ""


def format_date(date_num):
    s = str(date_num)
    if len(s) == 8:
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return s


def fetch_json(url):
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = resp.read().decode("utf-8")
            return json.loads(data)
    except Exception as e:
        print(f"  ❌ 请求失败: {e}")
        return None


def analyze_trades(data):
    """
    通过对比相邻交易日的持仓变化，推断买入和卖出行为
    """
    income_records = data["income"]
    trades_data = data["trades"]
    
    # 建立股票代码到名称的映射
    for date_str, trade_list in trades_data.items():
        for t in trade_list:
            code = stock_code_to_display(t.get("fullCode", ""))
            name = t.get("stockName", "")
            if code and name:
                STOCK_NAME_MAP[code] = name
    
    # 按日期排序（从早到晚）
    income_records.sort(key=lambda x: x["endDateNum"])
    
    # 提取每天的持仓股票列表（已去除前缀的代码）
    daily_holdings = {}
    daily_info = {}
    for r in income_records:
        date_num = r["endDateNum"]
        holdstocks = r.get("holdstocks") or ""
        if holdstocks:
            codes = set(stock_code_to_display(c) for c in holdstocks.split(","))
        else:
            codes = set()
        daily_holdings[date_num] = codes
        daily_info[date_num] = {
            "nowMoney": r.get("nowMoney", 0),
            "nowMoneyStr": r.get("nowMoneyStr", ""),
            "todayRateD": r.get("todayRateD", 0),
            "totalRateD": r.get("totalRateD", 0),
            "position": r.get("position", 0),
        }
    
    # 按时间顺序分析持仓变化
    dates = sorted(daily_holdings.keys())
    
    all_trades = []  # 所有推断的买卖记录
    
    # 跟踪每只股票的持仓起始日
    current_holdings = {}  # code -> buy_date
    
    for i, date in enumerate(dates):
        today_stocks = daily_holdings[date]
        
        if i == 0:
            # 第一天，所有股票都是"买入"
            for code in today_stocks:
                current_holdings[code] = date
                name = STOCK_NAME_MAP.get(code, "")
                all_trades.append({
                    "action": "买入",
                    "date": date,
                    "date_str": format_date(date),
                    "code": code,
                    "name": name,
                    "asset_after": daily_info[date]["nowMoneyStr"],
                    "day_return": daily_info[date]["todayRateD"],
                    "total_return": daily_info[date]["totalRateD"],
                    "position": daily_info[date]["position"],
                })
            continue
        
        prev_date = dates[i - 1]
        prev_stocks = daily_holdings[prev_date]
        
        # 新买入的股票 = 今天有但昨天没有
        new_buys = today_stocks - prev_stocks
        for code in new_buys:
            current_holdings[code] = date
            name = STOCK_NAME_MAP.get(code, "")
            all_trades.append({
                "action": "买入",
                "date": date,
                "date_str": format_date(date),
                "code": code,
                "name": name,
                "asset_after": daily_info[date]["nowMoneyStr"],
                "day_return": daily_info[date]["todayRateD"],
                "total_return": daily_info[date]["totalRateD"],
                "position": daily_info[date]["position"],
            })
        
        # 卖出的股票 = 昨天有但今天没有
        sold = prev_stocks - today_stocks
        for code in sold:
            buy_date = current_holdings.pop(code, None)
            name = STOCK_NAME_MAP.get(code, "")
            hold_days = 0
            if buy_date:
                # 计算持仓天数（交易日数）
                buy_idx = dates.index(buy_date) if buy_date in dates else -1
                sell_idx = i
                if buy_idx >= 0:
                    hold_days = sell_idx - buy_idx
            
            all_trades.append({
                "action": "卖出",
                "date": date,
                "date_str": format_date(date),
                "code": code,
                "name": name,
                "buy_date": format_date(buy_date) if buy_date else "",
                "hold_days": hold_days,
                "asset_after": daily_info[date]["nowMoneyStr"],
                "day_return": daily_info[date]["todayRateD"],
                "total_return": daily_info[date]["totalRateD"],
                "position": daily_info[date]["position"],
            })
    
    return all_trades


def download_images(data):
    """
    批量下载所有交割单截图
    """
    os.makedirs(IMAGES_DIR, exist_ok=True)
    
    income_records = data["income"]
    dates = sorted(set(str(r["endDateNum"]) for r in income_records))
    
    print(f"\n📸 开始下载交割单截图（共 {len(dates)} 个交易日）...")
    
    downloaded = 0
    failed = 0
    no_image = 0
    image_urls_map = {}
    
    for i, date_str in enumerate(dates):
        # 获取该日期的截图 URL
        url = f"https://www.tgb.cn/spmatch/gains/listUrl?spmatchSeq={SPMATCH_SEQ}&lookUserID={USER_ID}&dateNum={date_str}"
        result = fetch_json(url)
        
        if not result or not result.get("status"):
            failed += 1
            continue
        
        dto_list = result.get("dto", [])
        if not dto_list:
            no_image += 1
            continue
        
        for dto in dto_list:
            img_urls = dto.get("imgUrls", [])
            if not img_urls:
                no_image += 1
                continue
            
            image_urls_map[date_str] = img_urls
            
            for j, img_url in enumerate(img_urls):
                # 生成文件名
                suffix = f"_{j+1}" if len(img_urls) > 1 else ""
                ext = ".png" if img_url.endswith(".png") else ".jpg"
                filename = f"{date_str}{suffix}{ext}"
                filepath = os.path.join(IMAGES_DIR, filename)
                
                if os.path.exists(filepath):
                    downloaded += 1
                    continue
                
                try:
                    # 去掉 _sp760w 后缀获取原图
                    original_url = img_url.replace("_sp760w.png", ".jpg").replace("_sp760w", "")
                    req = urllib.request.Request(original_url, headers={
                        "User-Agent": "Mozilla/5.0",
                        "Referer": "https://www.tgb.cn/"
                    })
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        with open(filepath, "wb") as f:
                            f.write(resp.read())
                    downloaded += 1
                    if (downloaded % 20 == 0):
                        print(f"    📥 已下载 {downloaded} 张...")
                except Exception as e:
                    # 尝试用原始 URL
                    try:
                        req = urllib.request.Request(img_url, headers={
                            "User-Agent": "Mozilla/5.0",
                            "Referer": "https://www.tgb.cn/"
                        })
                        with urllib.request.urlopen(req, timeout=30) as resp:
                            with open(filepath, "wb") as f:
                                f.write(resp.read())
                        downloaded += 1
                    except Exception as e2:
                        print(f"    ❌ 下载失败 {date_str}: {e2}")
                        failed += 1
        
        if (i + 1) % 10 == 0:
            time.sleep(0.5)
        else:
            time.sleep(0.2)
    
    # 保存 URL 映射
    urls_json = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_截图URL.json")
    with open(urls_json, "w", encoding="utf-8") as f:
        json.dump(image_urls_map, f, ensure_ascii=False, indent=2)
    
    print(f"    ✅ 下载完成: {downloaded} 张, 无截图: {no_image} 天, 失败: {failed}")
    print(f"    📁 截图目录: {IMAGES_DIR}")
    print(f"    📁 URL映射: {urls_json}")
    
    return image_urls_map


def save_trades_data(all_trades):
    """保存买卖记录"""
    import csv
    
    # 保存 JSON
    with open(TRADES_JSON, "w", encoding="utf-8") as f:
        json.dump(all_trades, f, ensure_ascii=False, indent=2)
    
    # 保存 CSV
    with open(TRADES_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["操作", "日期", "股票代码", "股票名称", "买入日期", "持仓天数", 
                         "操作后总资产(万)", "当日收益(%)", "累计收益(%)", "仓位(%)"])
        for t in all_trades:
            writer.writerow([
                t.get("action"),
                t.get("date_str"),
                t.get("code"),
                t.get("name"),
                t.get("buy_date", ""),
                t.get("hold_days", ""),
                t.get("asset_after"),
                t.get("day_return"),
                t.get("total_return"),
                t.get("position"),
            ])
    
    print(f"    📁 买卖记录JSON: {TRADES_JSON}")
    print(f"    📁 买卖记录CSV: {TRADES_CSV}")


def save_excel(all_trades, image_urls_map, data):
    """生成包含买卖记录的 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("    ⚠️ openpyxl 不可用，跳过 Excel 生成")
        return
    
    output_excel = os.path.join(OUTPUT_DIR, "tgb_只核大学生_买卖记录.xlsx")
    
    wb = Workbook()
    
    # 样式
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10)
    green_font = Font(name="微软雅黑", size=10, color="008000", bold=True)
    red_font = Font(name="微软雅黑", size=10, color="FF0000", bold=True)
    buy_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    sell_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    title_font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="微软雅黑", size=10, color="666666")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # ============== Sheet 1: 买卖记录（按时间排序）==============
    ws1 = wb.active
    ws1.title = "买卖记录(时间序)"
    
    ws1.merge_cells('A1:J1')
    ws1['A1'] = "淘股吧 2025梦想杯 - 只核大学生 买卖记录（按时间排序）"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 35
    
    ws1.merge_cells('A2:J2')
    ws1['A2'] = "注：买卖时间通过对比每日持仓变化推断，价格信息请参考对应日期的交割单截图"
    ws1['A2'].font = subtitle_font
    ws1['A2'].alignment = center_align
    ws1.row_dimensions[2].height = 22
    
    headers = ["操作", "日期", "股票代码", "股票名称", "买入日期", "持仓天数(交易日)", 
               "操作后总资产(万)", "当日收益(%)", "累计收益(%)", "仓位(%)"]
    col_widths = [8, 14, 12, 14, 14, 16, 16, 13, 13, 10]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[3].height = 30
    
    # 按时间正序排列
    sorted_trades = sorted(all_trades, key=lambda x: (x["date"], x["action"]))
    
    for row_idx, t in enumerate(sorted_trades, 4):
        row_data = [
            t.get("action"),
            t.get("date_str"),
            t.get("code"),
            t.get("name"),
            t.get("buy_date", ""),
            t.get("hold_days", ""),
            t.get("asset_after"),
            t.get("day_return"),
            t.get("total_return"),
            t.get("position"),
        ]
        
        is_buy = t.get("action") == "买入"
        row_fill = buy_fill if is_buy else sell_fill
        action_font = green_font if is_buy else red_font
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = row_fill
            
            if col_idx == 1:
                cell.font = action_font
            else:
                cell.font = data_font
    
    ws1.freeze_panes = "A4"
    
    # ============== Sheet 2: 每笔交易汇总（配对买入卖出）==============
    ws2 = wb.create_sheet("完整交易(买卖配对)")
    
    ws2.merge_cells('A1:H1')
    ws2['A1'] = "淘股吧 2025梦想杯 - 只核大学生 完整交易记录（买卖配对）"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 35
    
    headers2 = ["序号", "股票代码", "股票名称", "买入日期", "卖出日期", "持仓天数(交易日)", 
                "买入日总资产(万)", "卖出日总资产(万)"]
    col_widths2 = [8, 12, 14, 14, 14, 16, 16, 16]
    
    for col_idx, (header, width) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[2].height = 30
    
    # 配对买卖记录
    buy_records = [t for t in all_trades if t["action"] == "买入"]
    sell_records = [t for t in all_trades if t["action"] == "卖出"]
    
    # 构建完整交易对
    paired_trades = []
    for sell in sell_records:
        paired_trades.append({
            "code": sell["code"],
            "name": sell["name"],
            "buy_date": sell.get("buy_date", ""),
            "sell_date": sell["date_str"],
            "hold_days": sell.get("hold_days", ""),
            "buy_asset": "",
            "sell_asset": sell.get("asset_after", ""),
        })
    
    # 找到未卖出的股票（最后一天仍持有的）
    last_date = max(t["date"] for t in all_trades)
    still_holding = set()
    for t in all_trades:
        if t["date"] == last_date and t["action"] == "买入":
            still_holding.add(t["code"])
    # 也检查持仓
    income_records = data["income"]
    income_records.sort(key=lambda x: x["endDateNum"], reverse=True)
    if income_records:
        last_hold = income_records[0].get("holdstocks") or ""
        if last_hold:
            for c in last_hold.split(","):
                code = stock_code_to_display(c)
                if code not in [s["code"] for s in sell_records if s["date"] == last_date]:
                    # 找这只股票最近的买入日期
                    buy_info = None
                    for b in reversed(buy_records):
                        if b["code"] == code:
                            buy_info = b
                            break
                    if buy_info:
                        paired_trades.append({
                            "code": code,
                            "name": STOCK_NAME_MAP.get(code, ""),
                            "buy_date": buy_info["date_str"],
                            "sell_date": "（仍持有）",
                            "hold_days": "",
                            "buy_asset": buy_info.get("asset_after", ""),
                            "sell_asset": "",
                        })
    
    # 按卖出日期排序
    paired_trades.sort(key=lambda x: x["sell_date"], reverse=True)
    
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    for row_idx, t in enumerate(paired_trades, 3):
        row_data = [
            row_idx - 2,
            t["code"],
            t["name"],
            t["buy_date"],
            t["sell_date"],
            t["hold_days"],
            t["buy_asset"],
            t["sell_asset"],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_align
            if (row_idx - 3) % 2 == 1:
                cell.fill = alt_fill
    
    ws2.freeze_panes = "A3"
    
    # ============== Sheet 3: 交割单截图URL ==============
    ws3 = wb.create_sheet("交割单截图URL")
    
    ws3.merge_cells('A1:C1')
    ws3['A1'] = "交割单截图URL索引（可点击查看原图，包含具体买卖价格）"
    ws3['A1'].font = title_font
    ws3['A1'].alignment = center_align
    ws3.row_dimensions[1].height = 35
    
    headers3 = ["日期", "截图URL", "本地文件"]
    col_widths3 = [14, 80, 40]
    
    for col_idx, (header, width) in enumerate(zip(headers3, col_widths3), 1):
        cell = ws3.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    ws3.row_dimensions[2].height = 30
    
    row_idx = 3
    for date_str in sorted(image_urls_map.keys(), reverse=True):
        urls = image_urls_map[date_str]
        for j, url in enumerate(urls):
            suffix = f"_{j+1}" if len(urls) > 1 else ""
            ext = ".png" if url.endswith(".png") else ".jpg"
            local_file = f"{date_str}{suffix}{ext}"
            
            ws3.cell(row=row_idx, column=1, value=format_date(date_str)).font = data_font
            ws3.cell(row=row_idx, column=1).border = thin_border
            ws3.cell(row=row_idx, column=1).alignment = center_align
            
            url_cell = ws3.cell(row=row_idx, column=2, value=url)
            url_cell.font = Font(name="微软雅黑", size=9, color="0563C1", underline="single")
            url_cell.hyperlink = url
            url_cell.border = thin_border
            
            ws3.cell(row=row_idx, column=3, value=local_file).font = data_font
            ws3.cell(row=row_idx, column=3).border = thin_border
            
            if (row_idx - 3) % 2 == 1:
                for c in range(1, 4):
                    ws3.cell(row=row_idx, column=c).fill = alt_fill
            
            row_idx += 1
    
    ws3.freeze_panes = "A3"
    
    # 保存
    wb.save(output_excel)
    print(f"\n    📁 Excel买卖记录: {output_excel}")
    print(f"       Sheet 1: 买卖记录(时间序) - {len(sorted_trades)} 条")
    print(f"       Sheet 2: 完整交易(买卖配对) - {len(paired_trades)} 条")
    print(f"       Sheet 3: 交割单截图URL - {row_idx - 3} 条")


def print_summary(all_trades):
    """打印交易统计摘要"""
    buys = [t for t in all_trades if t["action"] == "买入"]
    sells = [t for t in all_trades if t["action"] == "卖出"]
    
    print(f"\n📊 交易统计:")
    print(f"   买入次数: {len(buys)}")
    print(f"   卖出次数: {len(sells)}")
    
    # 持仓天数统计
    hold_days_list = [s["hold_days"] for s in sells if s.get("hold_days")]
    if hold_days_list:
        avg_hold = sum(hold_days_list) / len(hold_days_list)
        max_hold = max(hold_days_list)
        min_hold = min(hold_days_list)
        
        # 持仓分布
        t_plus_0 = sum(1 for d in hold_days_list if d == 0)
        t_plus_1 = sum(1 for d in hold_days_list if d == 1)
        t_plus_2_3 = sum(1 for d in hold_days_list if 2 <= d <= 3)
        t_plus_4_plus = sum(1 for d in hold_days_list if d >= 4)
        
        print(f"\n   持仓天数分析:")
        print(f"   平均持仓: {avg_hold:.1f} 个交易日")
        print(f"   最长持仓: {max_hold} 个交易日")
        print(f"   最短持仓: {min_hold} 个交易日")
        print(f"\n   持仓分布:")
        print(f"     T+0 (当日买卖): {t_plus_0} 次 ({t_plus_0/len(hold_days_list)*100:.1f}%)")
        print(f"     T+1 (隔日卖出): {t_plus_1} 次 ({t_plus_1/len(hold_days_list)*100:.1f}%)")
        print(f"     T+2~3:          {t_plus_2_3} 次 ({t_plus_2_3/len(hold_days_list)*100:.1f}%)")
        print(f"     T+4以上:        {t_plus_4_plus} 次 ({t_plus_4_plus/len(hold_days_list)*100:.1f}%)")
    
    # 最近 10 条交易
    recent = sorted(all_trades, key=lambda x: x["date"], reverse=True)[:15]
    print(f"\n📋 最近15条操作:")
    print(f"   {'操作':4s} {'日期':12s} {'代码':8s} {'名称':10s} {'买入日':12s} {'持仓':4s} {'总资产':8s}")
    print(f"   {'-'*70}")
    for t in recent:
        action = t["action"]
        date = t["date_str"]
        code = t["code"]
        name = t.get("name", "")[:8]
        buy_date = t.get("buy_date", "")
        hold = str(t.get("hold_days", ""))
        asset = t.get("asset_after", "")
        print(f"   {action:4s} {date:12s} {code:8s} {name:10s} {buy_date:12s} {hold:4s} {asset:8s}")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    print("=" * 60)
    print("📊 分析淘股吧交割单 - 推断买卖时间节点")
    print("=" * 60)
    
    # 读取已下载的数据
    print(f"\n📖 读取数据: {INPUT_JSON}")
    with open(INPUT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # 分析买卖行为
    print("\n🔍 分析持仓变化，推断买卖时间...")
    all_trades = analyze_trades(data)
    
    # 打印统计
    print_summary(all_trades)
    
    # 保存买卖记录
    print("\n💾 保存买卖记录...")
    save_trades_data(all_trades)
    
    # 下载交割单截图
    print("\n📸 下载交割单截图（包含实际买卖价格）...")
    image_urls_map = download_images(data)
    
    # 生成 Excel
    print("\n📝 生成 Excel 文件...")
    save_excel(all_trades, image_urls_map, data)
    
    print("\n" + "=" * 60)
    print("✅ 全部完成!")
    print(f"   💡 提示: 交割单截图中包含具体的买卖价格，")
    print(f"   请查看 {IMAGES_DIR} 目录下的图片")
    print("=" * 60)


if __name__ == "__main__":
    main()
