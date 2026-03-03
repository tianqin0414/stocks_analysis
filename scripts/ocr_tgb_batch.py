#!/usr/bin/env python3
"""
批量OCR提取6位高手交割单截图 - 适配多种券商截图格式
格式1(平安证券App): 忘忧阁主, 龙年大叔, 低调内敛的朋, 独行侠令狐冲, 只核大学生
格式2(网页表格): 天牌
"""
import json, os, re, sys, csv, time
from collections import defaultdict

import Vision, Quartz
from Foundation import NSURL

BASE_DIR = "/Users/tq/PycharmProjects/stocks_analysis"
SCREENSHOTS_DIR = os.path.join(BASE_DIR, "output/2_淘股吧高手/原始数据/截图")
OUTPUT_DIR = os.path.join(BASE_DIR, "output/2_淘股吧高手/原始数据")

KNOWN_STOCKS = set()

NOISE = {"持仓管理","持仓资讯","持仓分时","市值今","市值合","批量买入","批量卖出",
         "止盈止损","查看已清仓股票","首页","自选","行情","交易","资讯","买入","卖出",
         "撤单","持仓","查询","总资产","浮动盈亏","总市值","可用","可取","转账","国债",
         "淘股吧","淘股","淘限吧","淘服吧","淘险吧","淘胞吧","海吧","海限吧",
         "人民币","平安证券","国投证券","中航证券","华泰证券","招商证券","中信证券",
         "主线打板","仓位","当日参考盈亏","当日盈亏","参考盈亏","持仓/可用","成本/现价",
         "证券名称","证券代码","挂合管理","持仓/可用股吧","成本/现价合","成本/现价今",
         "盈亏","市值","盈亏成本价"}


def ocr_image(img_path):
    img_url = NSURL.fileURLWithPath_(img_path)
    src = Quartz.CGImageSourceCreateWithURL(img_url, None)
    if not src: return []
    cg = Quartz.CGImageSourceCreateImageAtIndex(src, 0, None)
    if not cg: return []
    req = Vision.VNRecognizeTextRequest.alloc().init()
    req.setRecognitionLanguages_(["zh-Hans", "en"])
    req.setRecognitionLevel_(Vision.VNRequestTextRecognitionLevelAccurate)
    handler = Vision.VNImageRequestHandler.alloc().initWithCGImage_options_(cg, {})
    ok, err = handler.performRequests_error_([req], None)
    if not ok: return []
    results = []
    for obs in req.results():
        t = obs.topCandidates_(1)[0].string()
        bb = obs.boundingBox()
        x, y = bb.origin.x, 1 - bb.origin.y - bb.size.height
        w, h = bb.size.width, bb.size.height
        results.append({"text":t, "x":x, "y":y, "w":w, "h":h, "xc":x+w/2, "yc":y+h/2})
    return results


def parse_num(text):
    if not text: return None
    t = text.strip().replace(",","").replace("，","").replace("$","").replace("¥","")
    t = re.sub(r'[淘股吧海限服胞险]', '', t).strip()
    t = re.sub(r'[\u4e00-\u9fff]+', '', t).strip()
    if not t or t == '--': return None
    # 处理百分号
    t = t.replace('%','').replace('％','')
    try: return float(t)
    except: return None


def is_stock(text):
    t = text.strip()
    if t in NOISE or len(t) < 2: return False
    if t in KNOWN_STOCKS: return True
    # 2~5个汉字
    if re.match(r'^[\u4e00-\u9fff]{2,5}$', t) and t not in NOISE: return True
    # ST股
    if re.match(r'^\*?ST[\u4e00-\u9fff]+$', t): return True
    # N+汉字 (新股)
    if re.match(r'^N[\u4e00-\u9fff]+$', t): return True
    # C+汉字
    if re.match(r'^C[\u4e00-\u9fff]+$', t): return True
    return False


def detect_format(ocr_results):
    """判断截图格式: app(手机App) 或 web(网页表格)"""
    texts = [r["text"] for r in ocr_results]
    full_text = " ".join(texts)
    # 网页格式特征: 有"证券化码"/"证券名券"/"股份金辣"等OCR误识别的表头
    # 或者有很长的列头行
    if "证券" in full_text and ("化码" in full_text or "名券" in full_text or "股份" in full_text):
        return "web"
    # 网页格式: 第一行就有"人民币：余額"
    if "余額" in full_text or "余额" in full_text[:50]:
        return "web"
    # App格式特征: 有"买入/卖出/撤单/持仓/查询"菜单
    menu_count = sum(1 for t in texts if t.strip() in ("买入","卖出","撤单","持仓","查询"))
    if menu_count >= 3:
        return "app"
    return "app"  # 默认


def extract_app(ocr_results):
    """提取App格式截图 (平安证券/国投证券/中航证券等)
    布局:
      左列(x<0.20): 股票名+市值
      中列(x 0.30-0.45): 盈亏金额+盈亏%
      持仓列(x 0.60-0.72): 持仓数量
      右列(x>0.80): 成本价/现价 (上下两行)
    """
    # 找持仓区域
    start_y = None
    end_y = 1.0
    for r in ocr_results:
        t = r["text"].strip()
        if t in ("市值","市值今","市值合") and r["x"] < 0.15:
            start_y = r["y"]
        if "查看已清仓" in t or "持仓管理" in t:
            if start_y and r["y"] > start_y:
                end_y = r["y"]; break

    if start_y is None:
        for r in ocr_results:
            if "持仓分时" in r["text"]: start_y = r["y"] + 0.03; break
    if start_y is None: start_y = 0.40

    # 提取总资产
    summary = {}
    for r in ocr_results:
        if "仓位" in r["text"]:
            m = re.search(r'([\d.]+)%', r["text"])
            if m: summary["仓位"] = float(m.group(1))
    for r in ocr_results:
        if r["x"] < 0.2 and 0.18 < r["y"] < start_y:
            v = parse_num(r["text"])
            if v and v > 10000: summary["总资产"] = v

    # 持仓区域
    items = [r for r in ocr_results if r["y"] > start_y + 0.02 and r["y"] < end_y]
    if not items: return [], summary

    # 找股票名 (左侧, x<0.20)
    stocks = []
    for r in items:
        if r["x"] < 0.20 and is_stock(r["text"].strip()):
            stocks.append({"name": r["text"].strip(), "y": r["y"]})
    stocks.sort(key=lambda x: x["y"])

    if not stocks: return [], summary

    holdings = []
    for i, s in enumerate(stocks):
        sy = s["y"]
        ny = stocks[i+1]["y"] if i+1 < len(stocks) else sy + 0.08
        row = [r for r in items if sy - 0.01 <= r["y"] <= ny - 0.005]

        h = {"股票名称": s["name"]}

        # 市值 (x<0.20, y>sy, 数字)
        for r in sorted(row, key=lambda x: x["y"]):
            if r["xc"] < 0.20 and r["y"] > sy + 0.005:
                v = parse_num(r["text"])
                if v is not None: h["市值"] = v; break

        # 成本/现价 (x>0.80, 两行数字)
        prices = []
        for r in sorted(row, key=lambda x: x["y"]):
            if r["xc"] > 0.80:
                v = parse_num(r["text"])
                if v is not None and v > 0: prices.append((r["y"], v))
        prices.sort(key=lambda x: x[0])
        if len(prices) >= 2:
            h["成本价"] = prices[0][1]
            h["现价"] = prices[1][1]
        elif len(prices) == 1:
            h["成本价"] = prices[0][1]

        # 盈亏% (中间列, 含%)
        for r in row:
            if 0.25 <= r["xc"] <= 0.50 and "%" in r["text"]:
                v = parse_num(r["text"])
                if v is not None: h["盈亏百分比"] = v; break

        # 盈亏金额
        for r in sorted(row, key=lambda x: x["y"]):
            if 0.25 <= r["xc"] <= 0.50 and "%" not in r["text"]:
                v = parse_num(r["text"])
                if v is not None: h["盈亏金额"] = v; break

        # 持仓数量 (x 0.55-0.75)
        for r in row:
            if 0.55 <= r["xc"] <= 0.75:
                v = parse_num(r["text"])
                if v is not None and v > 0: h["持仓数量"] = int(v); break

        # 交叉验证
        if h.get("成本价") and h.get("现价") and h["成本价"] > 0:
            h["盈亏百分比"] = round((h["现价"] - h["成本价"]) / h["成本价"] * 100, 2)

        KNOWN_STOCKS.add(s["name"])
        holdings.append(h)

    return holdings, summary


def extract_web(ocr_results):
    """提取网页表格格式截图 (天牌的券商网页版)
    布局: 一行式表格, 列头在y≈0.40
    证券代码 | 证券名称 | 股份金额 | 当前持仓 | ... | 盈亏成本价 | ... | 参考市值 | 价格 | 参考市值 | 浮动盈亏 | 盈亏比例
    股票数据在y≈0.60-0.80, 一只股票占两行(y差约0.04)
    """
    # 找到列头行
    header_y = None
    for r in ocr_results:
        if "证券" in r["text"] and len(r["text"]) > 20:
            header_y = r["y"]; break

    if header_y is None:
        for r in ocr_results:
            if r["text"].strip().startswith("证券") and r["y"] > 0.3:
                header_y = r["y"]; break

    if header_y is None: header_y = 0.40

    # 总资产
    summary = {}
    for r in ocr_results:
        if "总资产" in r["text"]:
            m = re.search(r'[\d,.]+', r["text"].split("总资产")[1] if "总资产" in r["text"] else "")
            if m:
                v = parse_num(m.group())
                if v: summary["总资产"] = v

    # 第一行的总资产
    top_items = [r for r in ocr_results if r["y"] < 0.10]
    for r in top_items:
        if "总资产" in r["text"]:
            m = re.search(r'总资产[：:]?\s*([\d,.]+)', r["text"])
            if m: summary["总资产"] = parse_num(m.group(1))

    # 数据行: 在header_y下方
    data_items = [r for r in ocr_results if r["y"] > header_y + 0.05]

    # 找股票代码+名称 (x<0.15, 格式如 "600580卧龙电驱")
    stocks = []
    for r in data_items:
        if r["x"] < 0.15:
            t = r["text"].strip()
            # 6位代码+名称
            m = re.match(r'^(\d{6})([\u4e00-\u9fff].+)$', t)
            if m:
                stocks.append({"code": m.group(1), "name": m.group(2), "y": r["y"]})
    
    if not stocks: return [], summary

    stocks.sort(key=lambda x: x["y"])

    holdings = []
    for i, s in enumerate(stocks):
        sy = s["y"]
        ny = stocks[i+1]["y"] if i+1 < len(stocks) else sy + 0.10
        row = [r for r in data_items if sy - 0.06 <= r["y"] <= ny - 0.01]

        h = {"股票名称": s["name"], "股票代码": s["code"]}

        # 收集该行的所有数字
        nums = []
        for r in sorted(row, key=lambda x: x["x"]):
            v = parse_num(r["text"])
            if v is not None:
                nums.append({"v": v, "x": r["x"], "xc": r["xc"], "y": r["y"], "text": r["text"]})

        # 天牌网页格式列位置 (根据OCR样本):
        # x≈0.17: 持仓数量(6000)
        # x≈0.43: 成本价(17.393)  
        # x≈0.55: 参考市值(104380.38)
        # x≈0.55: 现价(17.250) 和 参考市值(103500.00)
        # x≈0.72: 盈亏比例(-0.82)

        for n in nums:
            if 0.40 <= n["xc"] <= 0.52 and n["y"] < sy:
                h["成本价"] = n["v"]
            elif 0.52 <= n["xc"] <= 0.65 and n["y"] >= sy:
                if "现价" not in h:
                    h["现价"] = n["v"]
            elif 0.14 <= n["xc"] <= 0.25 and n["v"] > 0:
                h["持仓数量"] = int(n["v"])
            elif 0.70 <= n["xc"] <= 0.80:
                h["盈亏百分比"] = n["v"]

        # 如果没找到，从上面两行数字中猜
        if "成本价" not in h and "现价" not in h:
            price_nums = [n for n in nums if 0.40 <= n["xc"] <= 0.65 and 1 < n["v"] < 1000]
            price_nums.sort(key=lambda x: x["y"])
            if len(price_nums) >= 2:
                h["成本价"] = price_nums[0]["v"]
                h["现价"] = price_nums[1]["v"]

        if h.get("成本价") and h.get("现价") and h["成本价"] > 0:
            h["盈亏百分比"] = round((h["现价"] - h["成本价"]) / h["成本价"] * 100, 2)

        # 市值
        mv = [n for n in nums if n["v"] > 5000]
        if mv:
            h["市值"] = max(mv, key=lambda x: x["v"])["v"]

        KNOWN_STOCKS.add(s["name"])
        holdings.append(h)

    return holdings, summary


def process_master(name, img_dir):
    images = sorted([f for f in os.listdir(img_dir)
                     if (f.endswith('.png') or f.endswith('.jpg')) and re.search(r'\d{8}', f)])
    print(f"  📸 {len(images)}张截图", flush=True)

    all_records = []
    ok = 0; fail = 0; fmt_count = {"app":0, "web":0}

    for i, img_file in enumerate(images):
        m = re.search(r'(\d{8})', img_file)
        if not m: continue
        date_str = m.group(1)
        img_path = os.path.join(img_dir, img_file)

        try:
            ocr_results = ocr_image(img_path)
            if not ocr_results: fail += 1; continue

            fmt = detect_format(ocr_results)
            fmt_count[fmt] += 1

            if fmt == "web":
                holdings, summary = extract_web(ocr_results)
            else:
                holdings, summary = extract_app(ocr_results)

            if holdings:
                for h in holdings:
                    all_records.append({
                        "日期": f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}",
                        "高手名": name,
                        **h,
                        "总资产": summary.get("总资产"),
                        "仓位%": summary.get("仓位"),
                    })
                ok += 1
            else:
                fail += 1
        except Exception as e:
            fail += 1

        if (i+1) % 50 == 0:
            print(f"    [{i+1}/{len(images)}] ok={ok} fail={fail} records={len(all_records)} fmt={fmt_count}", flush=True)

    print(f"  ✅ ok={ok}/{len(images)} records={len(all_records)} fmt={fmt_count}", flush=True)
    return all_records


def save_csv(records, filepath):
    if not records: return
    cols = ["日期","高手名","股票名称","股票代码","成本价","现价","盈亏百分比","盈亏金额",
            "持仓数量","市值","总资产","仓位%"]
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
        w.writeheader()
        w.writerows(records)


def save_excel(all_records):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️ openpyxl不可用", flush=True); return

    wb = Workbook()
    hfont = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    dfont = Font(name="微软雅黑", size=10)
    gfont = Font(name="微软雅黑", size=10, color="008000")
    rfont = Font(name="微软雅黑", size=10, color="FF0000")
    border = Border(left=Side(style='thin',color='D9D9D9'), right=Side(style='thin',color='D9D9D9'),
                    top=Side(style='thin',color='D9D9D9'), bottom=Side(style='thin',color='D9D9D9'))
    alt = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    masters = sorted(set(r["高手名"] for r in all_records))
    for mi, master in enumerate(masters):
        ws = wb.active if mi == 0 else wb.create_sheet()
        ws.title = master[:10]
        recs = [r for r in all_records if r["高手名"] == master]
        headers = ["日期","股票名称","成本价","现价","盈亏%","持仓数量","市值","总资产"]
        widths = [14, 12, 12, 12, 10, 12, 14, 16]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri, r in enumerate(recs, 2):
            vals = [r.get("日期"), r.get("股票名称"), r.get("成本价"), r.get("现价"),
                    r.get("盈亏百分比"), r.get("持仓数量"), r.get("市值"), r.get("总资产")]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = dfont; cell.border = border
                if ri % 2 == 0: cell.fill = alt
                if ci == 5 and v is not None:
                    cell.font = gfont if v >= 0 else rfont

    out = os.path.join(OUTPUT_DIR, "全部高手_OCR持仓价格.xlsx")
    wb.save(out)
    print(f"  💾 Excel: {out}", flush=True)


def main():
    print("="*60, flush=True)
    print(f"📷 批量OCR (多格式适配) {time.strftime('%H:%M:%S')}", flush=True)
    print("="*60, flush=True)

    # 预加载股票名
    for fn in ["tgb_zhihedaxuesheng_data.json"]:
        fp = os.path.join(BASE_DIR, "output", fn)
        if os.path.exists(fp):
            with open(fp, encoding="utf-8") as f:
                for trades in json.load(f).get("trades",{}).values():
                    for t in trades:
                        if t.get("stockName"): KNOWN_STOCKS.add(t["stockName"])
    for bfile in os.listdir(os.path.join(BASE_DIR, "output/tgb_batch")):
        if bfile.endswith("_data.json"):
            fp = os.path.join(BASE_DIR, "output/tgb_batch", bfile)
            with open(fp, encoding="utf-8") as f:
                for trades in json.load(f).get("trades",{}).values():
                    for t in trades:
                        if t.get("stockName"): KNOWN_STOCKS.add(t["stockName"])
    print(f"  已知股票: {len(KNOWN_STOCKS)}个", flush=True)

    all_records = []
    masters = [
        ("天牌", "天牌_交割单截图"),
        ("忘忧阁主", "忘忧阁主_交割单截图"),
        ("低调内敛的朋", "低调内敛的朋_交割单截图"),
        ("独行侠令狐冲", "独行侠令狐冲_交割单截图"),
        ("龙年大叔", "龙年大叔_交割单截图"),
        ("只核大学生", "只核大学生_交割单截图"),
    ]

    for name, dirname in masters:
        img_dir = os.path.join(SCREENSHOTS_DIR, dirname)
        if not os.path.isdir(img_dir):
            print(f"\n⚠️ {name}: 目录不存在", flush=True); continue
        print(f"\n{'='*40}", flush=True)
        print(f"📥 {name}", flush=True)
        recs = process_master(name, img_dir)
        all_records.extend(recs)
        save_csv(recs, os.path.join(OUTPUT_DIR, f"{name}_OCR持仓数据.csv"))

    save_csv(all_records, os.path.join(OUTPUT_DIR, "全部高手_OCR持仓数据.csv"))
    print(f"\n💾 汇总CSV: 全部高手_OCR持仓数据.csv ({len(all_records)}条)", flush=True)

    save_excel(all_records)

    print(f"\n{'='*60}", flush=True)
    print(f"🎉 完成! {time.strftime('%H:%M:%S')}", flush=True)
    for name, _ in masters:
        n = sum(1 for r in all_records if r["高手名"] == name)
        cost_ok = sum(1 for r in all_records if r["高手名"] == name and r.get("成本价"))
        print(f"  {name}: {n}条 (有成本价{cost_ok}条)", flush=True)
    print(f"  总计: {len(all_records)}条", flush=True)

if __name__ == "__main__":
    main()
