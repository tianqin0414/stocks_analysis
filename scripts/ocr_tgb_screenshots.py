#!/usr/bin/env python3
"""
使用 macOS Vision 框架 OCR 提取淘股吧交割单截图中的持仓数据
提取每张截图中的：股票名称、市值、盈亏、持仓数量、成本价、现价
"""
import json
import os
import re
import sys
import csv
from collections import defaultdict

# macOS Vision framework
import Vision
import Quartz
from Foundation import NSURL

IMAGES_DIR = "/Users/tq/PycharmProjects/stocks_analysis/output/tgb_交割单截图"
OUTPUT_DIR = "/Users/tq/PycharmProjects/stocks_analysis/output"
OUTPUT_JSON = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_OCR持仓数据.json")
OUTPUT_CSV = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_OCR持仓数据.csv")

# 已知的股票名称列表（从之前下载的数据中获取，用于匹配）
# 这些名称用于识别截图中的持仓股行
KNOWN_STOCK_NAMES = set()


def ocr_image(img_path):
    """使用 macOS Vision 进行 OCR，返回按 y 坐标排序的文本行"""
    img_url = NSURL.fileURLWithPath_(img_path)
    img_source = Quartz.CGImageSourceCreateWithURL(img_url, None)
    if not img_source:
        return []
    
    cg_image = Quartz.CGImageSourceCreateImageAtIndex(img_source, 0, None)
    if not cg_image:
        return []
    
    request = Vision.VNRecognizeTextRequest.alloc().init()
    request.setRecognitionLanguages_(["zh-Hans", "en"])
    request.setRecognitionLevel_(Vision.VNRequestTextRecognitionLevelAccurate)
    
    handler = Vision.VNImageRequestHandler.alloc().initWithCGImage_options_(cg_image, {})
    success, error = handler.performRequests_error_([request], None)
    
    if not success:
        return []
    
    results = []
    for obs in request.results():
        text = obs.topCandidates_(1)[0].string()
        conf = obs.topCandidates_(1)[0].confidence()
        bbox = obs.boundingBox()
        # Convert from bottom-left origin to top-left origin
        x = bbox.origin.x
        y = 1 - bbox.origin.y - bbox.size.height
        w = bbox.size.width
        h = bbox.size.height
        results.append({
            "text": text,
            "confidence": conf,
            "x": x,
            "y": y,
            "w": w,
            "h": h,
            "x_center": x + w / 2,
            "y_center": y + h / 2,
        })
    
    return results


def is_stock_name(text):
    """判断是否为股票名称"""
    # 股票名称通常是2-4个中文字 + 可选后缀
    text = text.strip()
    # 排除已知的非股票名称
    non_stock = {"买入", "卖出", "撤单", "持仓", "查询", "首页", "行情", "自选", "交易",
                 "资讯", "理财", "总资产", "浮动盈亏", "总市值", "持仓股", "市值",
                 "盈亏", "可用", "可取", "国债", "转账", "查看已清仓股票",
                 "人民币账户A股", "中航证券", "挂合管理", "挂合答理",
                 "批量买入", "批量卖出", "止盈止损", "持仓/可用", "成本/现价",
                 "当日参考盈亏", "淘股吧", "淘股", "淘脂吧", "淘肪"}
    if text in non_stock:
        return False
    
    # Check if it's a known stock name
    if text in KNOWN_STOCK_NAMES:
        return True
    
    # Check pattern: 2-4 Chinese chars, possibly followed by stock type markers
    # 股票名称: 天创时尚, 顺灏股份, 航天动力, 广联航空, etc.
    if re.match(r'^[\u4e00-\u9fff]{2,5}$', text):
        # Additional check: should not be common words
        if text not in non_stock:
            return True
    
    # ST stocks
    if re.match(r'^[*]?ST[\u4e00-\u9fff]+$', text):
        return True
    
    return False


def parse_number(text):
    """解析数字字符串，处理逗号分隔符"""
    if not text:
        return None
    text = text.strip().replace(",", "").replace("，", "")
    # Remove any Chinese characters mixed in (OCR artifacts like 淘股)
    text = re.sub(r'[\u4e00-\u9fff]+', '', text)
    text = text.strip()
    if not text or text == '--':
        return None
    try:
        return float(text)
    except ValueError:
        return None


def extract_holdings_from_ocr(ocr_results, date_str):
    """
    从 OCR 结果中提取持仓数据
    
    截图布局（从左到右的列）:
    - 列1 (x~0.0-0.2): 股票名称 + 市值
    - 列2 (x~0.2-0.4): 盈亏金额 + 盈亏百分比  
    - 列3 (x~0.4-0.6): 持仓数量 / 可用数量
    - 列4 (x~0.6-0.8): 成本价 / 现价
    """
    if not ocr_results:
        return [], {}
    
    # 找到"持仓股"标记的 y 坐标，持仓数据在其下方
    holdings_start_y = None
    holdings_end_y = 1.0
    
    for item in ocr_results:
        text = item["text"].strip()
        if "持仓股" in text:
            holdings_start_y = item["y"]
        if "查看已清仓" in text or "挂合" in text or text == "首页":
            if holdings_start_y and item["y"] > holdings_start_y:
                holdings_end_y = item["y"]
                break
    
    if holdings_start_y is None:
        # 尝试找 "市值" 列头
        for item in ocr_results:
            if item["text"].strip() in ("市值", "市值令"):
                holdings_start_y = item["y"]
                break
    
    if holdings_start_y is None:
        holdings_start_y = 0.35  # 默认值
    
    # 提取总资产和仓位信息
    summary_info = {}
    for item in ocr_results:
        text = item["text"].strip()
        if "仓位" in text:
            match = re.search(r'([\d.]+)%', text)
            if match:
                summary_info["仓位"] = float(match.group(1))
        if "总资产" in text and item["y"] < holdings_start_y:
            summary_info["总资产_label_y"] = item["y"]
    
    # 查找总资产数值（在"总资产"标签附近）
    for item in ocr_results:
        text = item["text"].strip().replace(",", "")
        if item["x"] < 0.3 and item["y"] > 0.2 and item["y"] < holdings_start_y:
            val = parse_number(text)
            if val and val > 10000:
                summary_info["总资产"] = val
    
    # 在持仓区域中查找股票名称
    # 策略：先找到所有股票名称，然后对于每个名称，在同一行（相近的y坐标）找对应数据
    
    # 收集持仓区域的所有文本项，按 y 坐标排序
    holding_items = [item for item in ocr_results 
                     if item["y"] > holdings_start_y + 0.03 and item["y"] < holdings_end_y]
    
    if not holding_items:
        return [], summary_info
    
    # 找到所有股票名称及其 y 坐标
    stock_entries = []
    for item in holding_items:
        text = item["text"].strip()
        if is_stock_name(text) and item["x"] < 0.35:
            stock_entries.append({
                "name": text,
                "y": item["y"],
                "x": item["x"],
            })
    
    if not stock_entries:
        return [], summary_info
    
    # 按 y 坐标排序
    stock_entries.sort(key=lambda x: x["y"])
    
    # 对于每个股票名称，找到同一行的数据
    holdings = []
    
    for idx, stock in enumerate(stock_entries):
        stock_y = stock["y"]
        
        # 定义该股票的 y 范围（到下一个股票之间）
        if idx + 1 < len(stock_entries):
            next_y = stock_entries[idx + 1]["y"]
            y_range = (stock_y - 0.01, next_y - 0.005)
        else:
            y_range = (stock_y - 0.01, stock_y + 0.06)
        
        # 收集该行区域内的所有数据
        row_items = [item for item in holding_items 
                     if y_range[0] <= item["y"] <= y_range[1]]
        
        # 按 x 坐标和 y 坐标排序
        row_items.sort(key=lambda x: (x["x"], x["y"]))
        
        # 提取各列数据
        holding = {
            "股票名称": stock["name"],
            "市值": None,
            "盈亏金额": None,
            "盈亏百分比": None,
            "持仓数量": None,
            "可用数量": None,
            "成本价": None,
            "现价": None,
        }
        
        # 列分区 (x 坐标范围)
        #   列1: x < 0.25 (名称、市值)
        #   列2: x 0.25-0.50 (盈亏)
        #   列3: x 0.50-0.70 (持仓/可用)
        #   列4: x 0.70-1.0 (成本/现价)
        
        col1_items = [i for i in row_items if i["x_center"] < 0.28 and i["text"] != stock["name"]]
        col2_items = [i for i in row_items if 0.28 <= i["x_center"] < 0.52]
        col3_items = [i for i in row_items if 0.52 <= i["x_center"] < 0.72]
        col4_items = [i for i in row_items if i["x_center"] >= 0.72]
        
        # 列1: 市值（股票名称下方的数字）
        for item in col1_items:
            val = parse_number(item["text"])
            if val is not None and item["y"] > stock_y:
                holding["市值"] = val
                break
        
        # 列2: 盈亏金额和百分比
        col2_items.sort(key=lambda x: x["y"])
        for item in col2_items:
            text = item["text"].strip()
            if "%" in text:
                match = re.search(r'[-+]?[\d.]+%', text)
                if match:
                    val_str = match.group().replace("%", "")
                    try:
                        holding["盈亏百分比"] = float(val_str)
                    except ValueError:
                        pass
            else:
                val = parse_number(text)
                if val is not None:
                    if holding["盈亏金额"] is None:
                        # Check if it starts with - or +
                        clean_text = text.strip().replace(",", "").replace("，", "")
                        clean_text = re.sub(r'[\u4e00-\u9fff]+', '', clean_text).strip()
                        if clean_text.startswith('-') or clean_text.startswith('+'):
                            holding["盈亏金额"] = val
                        elif val < 100000:  # 盈亏一般不会太大
                            holding["盈亏金额"] = val
        
        # 列3: 持仓/可用
        col3_nums = []
        for item in col3_items:
            val = parse_number(item["text"])
            if val is not None:
                col3_nums.append((item["y"], int(val)))
        col3_nums.sort(key=lambda x: x[0])
        if len(col3_nums) >= 2:
            holding["持仓数量"] = col3_nums[0][1]
            holding["可用数量"] = col3_nums[1][1]
        elif len(col3_nums) == 1:
            holding["持仓数量"] = col3_nums[0][1]
        
        # 列4: 成本/现价
        col4_nums = []
        for item in col4_items:
            val = parse_number(item["text"])
            if val is not None:
                col4_nums.append((item["y"], val))
        col4_nums.sort(key=lambda x: x[0])
        if len(col4_nums) >= 2:
            holding["成本价"] = col4_nums[0][1]
            holding["现价"] = col4_nums[1][1]
        elif len(col4_nums) == 1:
            # 单个数字，可能是成本价或现价
            holding["成本价"] = col4_nums[0][1]
        
        # 判断是否已清仓
        if holding["市值"] == 0 and holding["持仓数量"] in (0, None):
            holding["状态"] = "已清仓"
        elif holding["持仓数量"] and holding["持仓数量"] > 0:
            holding["状态"] = "持仓中"
        else:
            holding["状态"] = "未知"
        
        holdings.append(holding)
    
    return holdings, summary_info


def load_known_stock_names():
    """从已下载数据中加载已知股票名称"""
    global KNOWN_STOCK_NAMES
    
    data_file = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_data.json")
    if os.path.exists(data_file):
        with open(data_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        for date_str, trade_list in data.get("trades", {}).items():
            for t in trade_list:
                name = t.get("stockName", "")
                if name:
                    KNOWN_STOCK_NAMES.add(name)
    
    print(f"  已加载 {len(KNOWN_STOCK_NAMES)} 个已知股票名称")


def main():
    print("=" * 60)
    print("📷 OCR 提取交割单截图中的持仓数据")
    print("=" * 60)
    
    # 加载已知股票名称
    load_known_stock_names()
    
    # 获取所有截图
    images = sorted([f for f in os.listdir(IMAGES_DIR) if f.endswith(('.png', '.jpg'))])
    print(f"\n📸 共 {len(images)} 张截图待处理")
    
    all_data = {}  # date -> {holdings, summary}
    failed = []
    
    for i, img_file in enumerate(images):
        date_str = img_file.split(".")[0].split("_")[0]  # 20251231
        img_path = os.path.join(IMAGES_DIR, img_file)
        
        if (i + 1) % 20 == 0 or i == 0:
            print(f"\n  [{i+1}/{len(images)}] 正在处理 {img_file}...")
        
        try:
            # OCR
            ocr_results = ocr_image(img_path)
            
            if not ocr_results:
                failed.append(date_str)
                continue
            
            # 提取持仓数据
            holdings, summary = extract_holdings_from_ocr(ocr_results, date_str)
            
            if holdings:
                all_data[date_str] = {
                    "holdings": holdings,
                    "summary": summary,
                }
                # 动态更新已知股票名称
                for h in holdings:
                    if h["股票名称"]:
                        KNOWN_STOCK_NAMES.add(h["股票名称"])
            else:
                failed.append(date_str)
        
        except Exception as e:
            print(f"    ❌ 处理失败 {img_file}: {e}")
            failed.append(date_str)
    
    print(f"\n✅ OCR 处理完成: {len(all_data)}/{len(images)} 张成功, {len(failed)} 张失败")
    
    if failed and len(failed) < 30:
        print(f"   失败日期: {', '.join(failed[:20])}")
    
    # ========== 保存 JSON ==========
    print(f"\n💾 保存 JSON: {OUTPUT_JSON}")
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    
    # ========== 保存 CSV ==========
    print(f"💾 保存 CSV: {OUTPUT_CSV}")
    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "日期", "股票名称", "状态", "市值", "盈亏金额", "盈亏百分比(%)",
            "持仓数量", "可用数量", "成本价", "现价", "仓位(%)", "总资产"
        ])
        
        for date_str in sorted(all_data.keys(), reverse=True):
            entry = all_data[date_str]
            summary = entry.get("summary", {})
            for h in entry["holdings"]:
                formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                writer.writerow([
                    formatted_date,
                    h["股票名称"],
                    h.get("状态", ""),
                    h["市值"],
                    h["盈亏金额"],
                    h["盈亏百分比"],
                    h["持仓数量"],
                    h["可用数量"],
                    h["成本价"],
                    h["现价"],
                    summary.get("仓位", ""),
                    summary.get("总资产", ""),
                ])
    
    # ========== 生成 Excel ==========
    generate_excel(all_data)
    
    # ========== 统计 ==========
    total_records = sum(len(entry["holdings"]) for entry in all_data.values())
    print(f"\n{'='*60}")
    print(f"✅ OCR 数据提取完成!")
    print(f"   📊 成功处理: {len(all_data)} 个交易日")
    print(f"   📊 持仓记录: {total_records} 条")
    print(f"   📊 失败: {len(failed)} 个")
    print(f"{'='*60}")


def generate_excel(all_data):
    """生成格式化的 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("   ⚠️ openpyxl 不可用")
        return
    
    output_excel = os.path.join(OUTPUT_DIR, "tgb_只核大学生_OCR持仓价格.xlsx")
    
    wb = Workbook()
    
    # 样式
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10)
    green_font = Font(name="微软雅黑", size=10, color="008000")
    red_font = Font(name="微软雅黑", size=10, color="FF0000")
    title_font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    hold_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    sold_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    
    # ========== Sheet 1: OCR 提取的持仓数据（按日期） ==========
    ws1 = wb.active
    ws1.title = "每日持仓价格(OCR)"
    
    ws1.merge_cells('A1:L1')
    ws1['A1'] = "淘股吧 2025梦想杯 - 只核大学生 每日持仓价格详情（OCR提取）"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 35
    
    headers = [
        "日期", "股票名称", "状态", "市值", "盈亏金额", "盈亏%",
        "持仓数量", "可用数量", "成本价", "现价", "仓位%", "总资产"
    ]
    col_widths = [14, 12, 8, 14, 14, 10, 12, 12, 12, 12, 10, 16]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[2].height = 30
    
    row_idx = 3
    for date_str in sorted(all_data.keys(), reverse=True):
        entry = all_data[date_str]
        summary = entry.get("summary", {})
        
        for h in entry["holdings"]:
            formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
            is_holding = h.get("状态") == "持仓中"
            
            row_data = [
                formatted_date,
                h["股票名称"],
                h.get("状态", ""),
                h["市值"],
                h["盈亏金额"],
                h["盈亏百分比"],
                h["持仓数量"],
                h["可用数量"],
                h["成本价"],
                h["现价"],
                summary.get("仓位", ""),
                summary.get("总资产", ""),
            ]
            
            row_bg = hold_fill if is_holding else sold_fill if h.get("状态") == "已清仓" else None
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.font = data_font
                
                if col_idx in (1, 2, 3):
                    cell.alignment = center_align
                else:
                    cell.alignment = right_align
                
                # 盈亏颜色
                if col_idx in (5, 6) and val is not None:
                    try:
                        v = float(val)
                        cell.font = green_font if v > 0 else red_font if v < 0 else data_font
                    except (ValueError, TypeError):
                        pass
                
                if row_bg:
                    cell.fill = row_bg
            
            row_idx += 1
    
    ws1.freeze_panes = "A3"
    
    # ========== Sheet 2: 与买卖记录合并 ==========
    # 读取之前推断的买卖记录
    trades_json = os.path.join(OUTPUT_DIR, "tgb_zhihedaxuesheng_买卖记录.json")
    if os.path.exists(trades_json):
        with open(trades_json, "r", encoding="utf-8") as f:
            trade_records = json.load(f)
        
        ws2 = wb.create_sheet("买卖价格合并")
        
        ws2.merge_cells('A1:L1')
        ws2['A1'] = "买卖记录 + OCR价格数据合并"
        ws2['A1'].font = title_font
        ws2['A1'].alignment = center_align
        ws2.row_dimensions[1].height = 35
        
        headers2 = [
            "操作", "日期", "股票代码", "股票名称", "买入日期", "持仓天数",
            "成本价(OCR)", "现价(OCR)", "持仓数量(OCR)", "市值(OCR)", "盈亏(OCR)", "当日总资产(万)"
        ]
        col_widths2 = [8, 14, 12, 12, 14, 10, 12, 12, 12, 14, 14, 14]
        
        for col_idx, (header, width) in enumerate(zip(headers2, col_widths2), 1):
            cell = ws2.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws2.column_dimensions[get_column_letter(col_idx)].width = width
        ws2.row_dimensions[2].height = 30
        
        # 构建 OCR 数据索引 (date, stock_name) -> holding info
        ocr_index = {}
        for date_str, entry in all_data.items():
            for h in entry["holdings"]:
                key = (date_str, h["股票名称"])
                ocr_index[key] = h
        
        row_idx = 3
        # 按时间倒序
        sorted_trades = sorted(trade_records, key=lambda x: x["date"], reverse=True)
        
        for t in sorted_trades:
            date_str = str(t["date"])
            stock_name = t.get("name", "")
            
            # 查找 OCR 数据
            ocr_data = ocr_index.get((date_str, stock_name), {})
            
            is_buy = t["action"] == "买入"
            buy_fill_style = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            sell_fill_style = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            
            row_data = [
                t["action"],
                t["date_str"],
                t.get("code", ""),
                stock_name,
                t.get("buy_date", ""),
                t.get("hold_days", ""),
                ocr_data.get("成本价"),
                ocr_data.get("现价"),
                ocr_data.get("持仓数量"),
                ocr_data.get("市值"),
                ocr_data.get("盈亏金额"),
                t.get("asset_after", ""),
            ]
            
            row_fill = buy_fill_style if is_buy else sell_fill_style
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = center_align if col_idx <= 6 else right_align
                cell.fill = row_fill
                
                if col_idx == 1:
                    cell.font = Font(name="微软雅黑", size=10, bold=True,
                                    color="008000" if is_buy else "FF0000")
                else:
                    cell.font = data_font
            
            row_idx += 1
        
        ws2.freeze_panes = "A3"
    
    # 保存
    wb.save(output_excel)
    print(f"\n   📁 Excel(OCR价格): {output_excel}")


if __name__ == "__main__":
    main()
